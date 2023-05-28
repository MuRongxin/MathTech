from openpyxl import*
from openpyxl.styles import*
import os
import win32com.client as win32
import re
import math
# from tqdm import tqdm

# xmlFileName=input("原始成绩文件名: ")
# targetFileName=input("目标文件名：")
# targetSheetFocusName=input("目标表名(40 or 100): ")
# fullScore=int(input("理论满分："))


# 获取当前所在的目录
current_dir = os.path.dirname(os.path.abspath(__file__))
# 列出当前目录下的所有。xls文件
files = [f for f in os.listdir(current_dir) if os.path.isfile(os.path.join(current_dir, f))and f.endswith('.xls')]

targetSheetName="Alldata"

targetFileName=None
targetSheetFocusName=None
fullScore=None

def changeXlsToXlsx(fileName):    
    if os.path.isfile(fileName.split('.')[0]+".xlsx"):
        print("文件存在")
        return
    fileName = os.getcwd()+"\\"+fileName
    Excelapp = win32.gencache.EnsureDispatch('Excel.Application')
    workbook = Excelapp.Workbooks.Open(fileName)
    # 转xlsx时: FileFormat=51,
    # 转xls时:  FileFormat=56,
    workbook.SaveAs(fileName.replace('xls', 'xlsx'), FileFormat=51)
    workbook.Close()
    Excelapp.Application.Quit()




#changeXlsToXlsx(xmlFileName)

workBookTar=None

workSheetOri = None
workSheetTar=None
workFocusSheetTar=None
# targetSheetFocusName= None

def OpenTableSheet(oriXlsxFile,tarXlsxFile,targetSheetFocusName):
    workBookOri=load_workbook(oriXlsxFile+".xlsx")#打开源文件
    # workBookOri=load_workbook(oriXlsxFileName)#打开源文件
    global workSheetOri
    workSheetOri=workBookOri.active

    #workBookTar=load_workbook(targetFileName)#打开目标文件
    global workBookTar
    
    workBookTar=load_workbook(tarXlsxFile)#打开目标文件
    global workSheetTar
    workSheetTar=workBookTar[targetSheetName]#打开成绩主表

    global workFocusSheetTar
    workFocusSheetTar=workBookTar[targetSheetFocusName]#打开占比表

nameCol=0 #标记姓名列
scoreCol=0#标记成绩列

score=[]#成绩列表

def GetOriSCore(): 
    global nameCol
    global scoreCol
    global score
    score.clear()   
    for i in range(1, workSheetOri.max_column+1):
        data=workSheetOri.cell(row=1, column=i)
        if data.value=="姓名":
            nameCol=data.column
        if data.value=="成绩":
            scoreCol=data.column

    for i in range(2, workSheetOri.max_row+1):
        nameData=workSheetOri.cell(row=i, column=nameCol)
        scoreData=workSheetOri.cell(row=i, column=scoreCol)
        if nameData.value=="最高分" or nameData.value=="最低分" or nameData.value=="平均分" or nameData.value=="及格率" or nameData.value=="优秀率" or nameData.value=="低分率":
            continue
        score.append( [str(nameData.value),str(scoreData.value)])
        #print(nameData.value)


tarCol=None
tarRow=None
tarCol_focus=None
tarRow_focus=None
date=None
dateList=[]

def MarkIndex():
    global tarCol 
    global tarRow 
    global tarCol_focus
    global tarRow_focus

    tarCol=workSheetTar.max_column
    tarRow=workSheetTar.max_row

    tarCol_focus=workFocusSheetTar.max_column
    tarRow_focus=workFocusSheetTar.max_row

    # date=time.strftime('%Y-%m-%d',time.localtime(time.time()))

# isDate=input("是否采用系统时间 [y/n]: ")

# if isDate == 'n':
#     date=input("请输入时间: ")

alignment_center = Alignment(horizontal='center', vertical='center')#定义居中对齐
alignment_right=Alignment(horizontal='right', vertical='center')#定义右对齐

isCalculate=False
isReadDate=False

def SetBasicFormat(date):
    global isReadDate
    if isReadDate == False:
        for i in range(1,tarCol+1):
            dateList.append(workSheetTar.cell(row=1, column=i).value)
        isReadDate=True;
    
    if date in dateList:
        print("该数据已经处理")
        global isCalculate
        isCalculate=True
        return
    #设置日期
    workSheetTar.cell(row=1, column=tarCol+1).value=date
    workFocusSheetTar.cell(row=1, column=tarCol_focus+1).value=date

    workSheetTar.cell(row=1, column=tarCol+1).alignment=alignment_center
    workSheetTar.merge_cells(start_row=1, end_row=1, start_column=tarCol+1, end_column=tarCol+3)

    workSheetTar.cell(row=2, column=tarCol+1).value="理论满分"
    workSheetTar.cell(row=2, column=tarCol+2).value="原始成绩"
    workSheetTar.cell(row=2, column=tarCol+3).value="成绩占比"

    isCalculate=False

mean=0
variance=0
std_dev=0

scores=[]

def calculate_mean_and_variance(data):
    # 计算平均值
    global mean
    mean = sum(data) / len(data)    
    # 计算方差
    global variance
    variance = sum([(x - mean)**2 for x in data]) / len(data)
    # 计算标准差
    global std_dev
    std_dev = math.sqrt(sum([(x - mean)**2 for x in data]) / len(data))    
   
def GetScores():
    scores.clear()
    for i in score:
        scores.append(int(i[1]))


def standardize_scores(data):
    # 计算数据集的平均值和标准差
    # mean = sum(data) / len(data)
    # std_dev = (sum([(x - mean)**2 for x in data]) / len(data)) ** 0.5    
    # 对每个分数进行标准化处理
    # standardized_data = [(x - mean) / std_dev for x in data]
    
    z = (data - mean) / std_dev

    return z


def SetScore():

    GetScores()
    
    calculate_mean_and_variance(scores)    

    for i in range(3,tarRow+1):
        workSheetTar.cell(row=i, column=tarCol+1).value=fullScore
        workSheetTar.cell(row=i, column=tarCol+1).alignment=alignment_right
        name=workSheetTar.cell(row=i, column=2).value    
        for j in range(0,len(score)-1):
            student = score[j]        
            if student[0]==name:            
                workSheetTar.cell(row=i, column=tarCol+2).value=int(student[1])
                workSheetTar.cell(row=i, column=tarCol+2).alignment=alignment_right
                # workSheetTar.cell(row=i, column=tarCol+3).value=int(student[1])/fullScore
                workSheetTar.cell(row=i, column=tarCol+3).value=standardize_scores(int(student[1]))
                workSheetTar.cell(row=i, column=tarCol+3).alignment=alignment_right 

                for x in range(2,tarRow_focus+1):  
                    if workFocusSheetTar.cell(row=x, column=tarCol_focus+1).value==None:
                        workFocusSheetTar.cell(row=x, column=tarCol_focus+1).value=standardize_scores(0)            
                    if workFocusSheetTar.cell(row=x, column=1).value==name:
                        # workFocusSheetTar.cell(row=x, column=tarCol_focus+1).value=int(student[1])/fullScore                        
                        workFocusSheetTar.cell(row=x, column=tarCol_focus+1).value=standardize_scores(int(student[1]))
                        # break
                    # else:
                    #     workFocusSheetTar.cell(row=x, column=tarCol_focus+1).value=int(0)/fullScore

                del(score[j])
                # break
                pass
            
    # print("当前成绩列表剩余: ")
    # for i in score:
    #     print(i)
   
              

# 打印出所有文件名 日期_[次数]_分数_班级
for f in files:    
        
    changeXlsToXlsx(f)
    
    oriFileName = f.split('.')[0]
    fileNameCompositionList=re.findall('\d+', f)
    length=len(fileNameCompositionList)    
    
    if(fileNameCompositionList[length-1]=="142"):
        targetFileName="Exam142Score.xlsx"
    if(fileNameCompositionList[length-1]=="145"):
        targetFileName="Exam145Score.xlsx"

    if(int(fileNameCompositionList[length-2])>60):
        fullScore=int(fileNameCompositionList[length-2])
        targetSheetFocusName="100"
    else:
        fullScore=int(fileNameCompositionList[length-2])
        targetSheetFocusName="40"

    date = fileNameCompositionList[0]
    date = "2023/"+date[0:2]+"/"+date[2:4]  
    # date = "2023."+date[0:2]+"."+date[2:4]  

    if length==4:
        date+="_"+fileNameCompositionList[1]
    print("当前文件："+str(oriFileName)+" 目标文件："+str(targetFileName)+" 目标占比表："+str(targetSheetFocusName))
       
    OpenTableSheet(oriFileName,targetFileName,targetSheetFocusName)

    GetOriSCore()

    MarkIndex()
    
    SetBasicFormat(date)
   
    if isCalculate:
        continue
    
    SetScore()
    print("目前均值："+ str(mean))

    #print(re.findall('\d+', f))

    workBookTar.save(targetFileName)




# print(workSheetTar.cell(row=workSheetTar.max_row, column=workSheetTar.max_column).value)

# print("当前行数："+str(workSheetTar.max_row)+"当前列数："+str(workSheetTar.max_column))
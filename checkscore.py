import os
import openpyxl
import sys
from openpyxl.styles import Alignment
from openpyxl.styles.colors import Color
from openpyxl.styles import Font, NamedStyle,PatternFill



souseExcelFile=r"../file/2023年12月高一第四次月考成绩汇总分析表核对版.xlsx"

# if len(sys.argv)>1:
#     # resoultclassName=sys.argv[1]
#     print("canshu 1")
# elif len(sys.argv)>2:
#     # souseExcelFile=sys.argv[2]
#     print("canshu 2")
# else:
#     # resoultclassName="150"
#     print("canshu 0")

resoultclassName="153"
# print(resoultclassName)
wb = openpyxl.load_workbook(souseExcelFile)
# print(souseExcelFile)

# wb = openpyxl.load_workbook(souseExcelFile)
#遍历所有的sheet

sheet=wb.sheetnames

sourcesheet_1=wb["物化生总表"]
sourcesheet_2=wb["政史地总表"]

#创建目标文件
resoultExcelFile=openpyxl.Workbook()
resoultExcelFile.remove(resoultExcelFile.active)
resoultSheet=resoultExcelFile.create_sheet(title="原"+resoultclassName+"数学得分明细")
resoultAllSheet=resoultExcelFile.create_sheet(title="原"+resoultclassName+"全科目得分")
resoultAllSheet_2=resoultExcelFile.create_sheet(title="原"+resoultclassName+"全科目得分文")

merged_ranges = sourcesheet_1.merged_cells
merged_ranges_math = sourcesheet_1.merged_cells

for merged_range in merged_ranges:
    if merged_range.max_row>100:
        merged_ranges.remove(merged_range)    
for merged_range in merged_ranges:    
    resoultAllSheet.merge_cells(str(merged_range))

baserow=0
def setFotmat(sheet,row,col,font,fill):
    for i in range(1,sourcesheet_1.max_row+1):
        for j in range(3,7):    
            if isinstance(sourcesheet_1.cell(row=i,column=j).value, (int, float)):
                if j==6:
                    baserow=i
                    break          
            else:
                break
        if baserow>0:
            break
    for i in range(1,sourcesheet_1.max_column+1):
        for j in range(1,baserow):   
            resoultAllSheet.cell(row=j,column=i,value=sourcesheet_1.cell(row=j,column=i).value)
        # resoultAllSheet.cell(row=2,column=i,value=sourcesheet.cell(row=2,column=i).value)

for i in range(1,sourcesheet_1.max_column+1):
    resoultAllSheet.cell(row=1,column=i,value=sourcesheet_1.cell(row=1,column=i).value)
    resoultAllSheet.cell(row=2,column=i,value=sourcesheet_1.cell(row=2,column=i).value)

for i in range(1,sourcesheet_2.max_column+1):
    resoultAllSheet_2.cell(row=1,column=i,value=sourcesheet_1.cell(row=1,column=i).value)
    resoultAllSheet_2.cell(row=2,column=i,value=sourcesheet_1.cell(row=2,column=i).value)

# index=0
# for row in sourcesheet_2.iter_rows(min_row=2, max_row=2):
#     for cell in row:
#         index+=1
#         print(str(index), end=' ')
#         print(cell.value, end=' ')
#     print()
indexz=18
for i in range(resoultAllSheet.max_column+1,resoultAllSheet.max_column+sourcesheet_2.max_column+1):
    resoultAllSheet.cell(row=2,column=i ,value=sourcesheet_2.cell(row=2,column=indexz).value)
    indexz+=1
    # resoultSheet.cell(row=2,column=i,value=sourceWorkSheet.cell(row=2,column=i).value)#要注意调整这个东西
  

with open("data"+resoultclassName+".txt", 'r',encoding='utf-8') as f:
    lines = f.readlines()
# 提取姓名列表
names = []
for line in lines:
    name = line.split(' ')[1].strip()  # 使用split()方法分割文本
    names.append(name)

max_row = 0
max_column = 0
namemark=0
mathmark=0



for row in sourcesheet_1.iter_rows(): # 遍历工作表的每一列
    for cell in row:
        # 遍历单元格的值
        if cell.value =='姓名':
            namemark = cell.column               
        if cell.value =='数学':
            mathmark=cell.column
            break
    if mathmark>0:
        break

max_column=sourcesheet_1.max_column
max_row=sourcesheet_1.max_row


for i in range(1,max_row+1):
    resoultAllSheetRow=resoultAllSheet.max_row+1
    for j in range(0,len(names)):
        if sourcesheet_1.cell(row=i,column=namemark).value ==names[j] and names:
            # names.remove(names[j])#避免重复
            for z in range(1,max_column+1):
                resoultAllSheet.cell(row=resoultAllSheetRow,column=z).value=sourcesheet_1.cell(row=i,column=z).value
            
        # if(i>sourcesheet_2.max_row):
        #     continue
        # resoultAllSheetRow=resoultAllSheet.max_row+1
        # if sourcesheet_2.cell(row=i,column=namemark).value ==names[j]:
        #     # names.remove(names[j])#避免重复
        #     for z in range(1,max_column+1):
        #         resoultAllSheet.cell(row=resoultAllSheetRow,column=z).value=sourcesheet_2.cell(row=i,column=z).value
for i in range(1,sourcesheet_2.max_row+1):
    resoultAllSheetRow=resoultAllSheet_2.max_row+1
    for j in range(0,len(names)):
        if sourcesheet_2.cell(row=i,column=namemark).value ==names[j] and names:
            # names.remove(names[j])#避免重复
            for z in range(1,max_column+1):
                resoultAllSheet_2.cell(row=resoultAllSheetRow,column=z).value=sourcesheet_2.cell(row=i,column=z).value
    

#TODO

resoultExcelFile.save('result'+resoultclassName+'.xlsx')

def solutionSheet(sheet):
    # 遍历工作表的每一行
    for row in sheet.iter_rows():
    # 遍历工作表的每一列
        for cell in row:
        # 遍历单元格的值
            if cell.value =='姓名':
                namemark = cell.column               
            if cell.value =='数学':
                mathmark=cell.column
                break
        break
    # breakpoint()
    for i in range(1,max_row+1):
        for j in range(0,len(names)):
            if sheet.cell(row=i,column=namemark).value ==names[j]:
                names.remove(names[j])#避免重复
                for z in range(0,max_column+1):
                    resoultAllSheet.cell(row=resoultAllSheet.max_row+1 ,column=z).value=sheet.cell(row=i,column=z).value
#TODO
# 遍历所有工作表
# for sheet_name in wb.sheetnames:
#   sheet = wb[sheet_name]  # 获取工作表
#   print(f'Working on sheet: {sheet_name}')
#   # 在这里处理工作表数据...
#   max_row = sheet.max_row
#   max_column = sheet.max_column
#   solutionSheet(sheet)


# 创建一个加粗字体的样式
bold_font = Font(bold=True)
bold_style = NamedStyle(name='bold_style')
bold_style.font = bold_font

# 创建一个填充颜色为红色的PatternFill对象
fill_color = Color(rgb='39c5bb')
fill = PatternFill(fill_type='solid', start_color=fill_color, end_color=fill_color)

fill_color_2 = Color(rgb='dcdcdc')
fill_2 = PatternFill(fill_type='solid', start_color=fill_color_2, end_color=fill_color_2)

fill_color_3 = Color(rgb='7fffd4')
fill_3 = PatternFill(fill_type='solid', start_color=fill_color_3, end_color=fill_color_3)





resoultExcelFile.save("2023年12月考试 原"+resoultclassName+"考试成绩.xlsx")



import zipfile
import os
import win32com.client as win32
import openpyxl as xl

resoultExcelFile=xl.Workbook()
resoultSheet = resoultExcelFile.active


sourceAllExcelFile=xl.load_workbook("【2023年5月高一月考】全部考生成绩汇总.xlsx")
sourceAllWorkSheet = sourceAllExcelFile.active
merged_ranges = sourceAllWorkSheet.merged_cells

def parser_merged_cell(sheet, row, col):
    # 获取指定单元格对象
    cell = sheet.cell(row=row, column=col)
    # 判断该单元格是否为合并单元格
    if isinstance(cell, xl.cell.MergedCell):
        # 如果是，循环查找该单元格所属的合并区域
        for merged_range in merged_ranges:
            # 如果该单元格坐标在合并区域内
            if cell.coordinate in merged_range:
                # 获取合并区域左上角的单元格作为该单元格的值返回
                cell = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                break
    # 返回该单元格的值
    return cell.value

# 调用函数，传入要读取的单元格的行和列

# print(sourceAllWorkSheet.cell(row=3,column=7))

# 遍历合并区域列表
for merged_range in merged_ranges:
    # 打印合并区域的坐标
    print(merged_range.min_row, merged_range.max_row, merged_range.min_col, merged_range.max_col)
    # 打印合并区域的字符串表示
    # resoultSheet.merge_cells(str(merged_range))
    print(merged_range)
    if merged_range.max_row>100:
        merged_ranges.remove(merged_range)
print("=========")
for merged_range in merged_ranges:
    # 打印合并区域的坐标
    print(merged_range.min_row, merged_range.max_row, merged_range.min_col, merged_range.max_col)
    # 打印合并区域的字符串表示
    resoultSheet.merge_cells(str(merged_range))
    print(merged_range)

max_row_all = sourceAllWorkSheet.max_row
max_column_all = sourceAllWorkSheet.max_column
for i in range(1,max_column_all+1):
    resoultSheet.cell(row=1,column=i,value=sourceAllWorkSheet.cell(row=1,column=i).value)
    resoultSheet.cell(row=2,column=i,value=sourceAllWorkSheet.cell(row=2,column=i).value)
    resoultSheet.cell(row=3,column=i,value=sourceAllWorkSheet.cell(row=3,column=i).value)


# value = parser_merged_cell(sourceAllWorkSheet, 3, 7)
# print(value)

resoultExcelFile.save("Temp.xlsx")

# def zipdir(path, zipname):
#     with zipfile.ZipFile(zipname, 'w', zipfile.ZIP_DEFLATED) as ziph:
#         # 遍历文件夹中的所有文件
#         for root, dirs, files in os.walk(path):
#             for file in files:
#                 # 获取文件的绝对路径
#                 filepath = os.path.join(root, file)
#                 # 将文件添加到zip文件中
#                 ziph.write(filepath)

# # 调用zipdir函数打包文件夹
# folder_path = r'C:\Users\PC\Downloads\slicer-gui-windows-v1.1.0\lrin'
# zip_name = 'lrin.zip'
# zipdir(folder_path, zip_name)
# # 帮我写一份情书吧，我们都是高中生

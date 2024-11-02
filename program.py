import os
import openpyxl
from openpyxl.styles.colors import Color
from tqdm import tqdm
from openpyxl.styles import Alignment
from openpyxl.styles import Font, NamedStyle,PatternFill


#得到目标文件
sourceExcelFile=openpyxl.load_workbook("【祥华2023年6月高一月考联考】所有班级学生小题得分明细.xlsx")
sourceAllExcelFile=openpyxl.load_workbook("祥华2023年6月高一月考联考-学生成绩（全部考生）.xlsx")

resoultExcelFile=openpyxl.Workbook()
resoultExcelFile.remove(resoultExcelFile.active)

sourceWorkSheet = sourceExcelFile["数学"]
sourceAllWorkSheet = sourceAllExcelFile.active

className = "143"

resoultSheet=resoultExcelFile.create_sheet(title="原"+className+"数学得分明细")
resoultAllSheet=resoultExcelFile.create_sheet(title="原"+className+"全科目得分")

# 获取最大行和最大列
max_row = sourceWorkSheet.max_row
max_column = sourceWorkSheet.max_column

max_row_all = sourceAllWorkSheet.max_row
max_column_all = sourceAllWorkSheet.max_column

markRow_name=0
markRow_all_name=0
markRow_all_math=0
markRow_all_allScore=0
markRow_art=0

merged_ranges = sourceAllWorkSheet.merged_cells
merged_ranges_math = sourceWorkSheet.merged_cells

# 将目标表合并
for merged_range in merged_ranges:
    if merged_range.max_row>100:
        merged_ranges.remove(merged_range)    
for merged_range in merged_ranges:    
    resoultAllSheet.merge_cells(str(merged_range))

# for merged_range in merged_ranges_math:
#     if merged_range.max_row>100:
#         merged_ranges_math.remove(merged_range)    
# for merged_range in merged_ranges_math:    
#     resoultSheet.merge_cells(str(merged_range))

#写入表头
for i in range(1,max_column+1):
    resoultSheet.cell(row=1,column=i,value=sourceWorkSheet.cell(row=1,column=i).value)
    resoultSheet.cell(row=2,column=i,value=sourceWorkSheet.cell(row=2,column=i).value)
    # breakpoint()
    if sourceWorkSheet.cell(row=2,column=i).value=="姓名":
        markRow_name=i
# markRow_name=2   
for i in range(1,max_column_all+1):
    resoultAllSheet.cell(row=1,column=i,value=sourceAllWorkSheet.cell(row=1,column=i).value)
    resoultAllSheet.cell(row=2,column=i,value=sourceAllWorkSheet.cell(row=2,column=i).value)
    # resoultAllSheet.cell(row=3,column=i,value=sourceAllWorkSheet.cell(row=3,column=i).value)

    if sourceAllWorkSheet.cell(row=2,column=i).value=="姓名":
        markRow_all_name=i
    if sourceAllWorkSheet.cell(row=2,column=i).value=="数学":
        markRow_all_math=i
    if sourceAllWorkSheet.cell(row=2,column=i).value=="总分": #要注意调整这个东西
        markRow_all_allScore=i
    if sourceAllWorkSheet.cell(row=2,column=i).value=="地理": #要注意调整这个东西
        markRow_art=i

with open("data"+className+".txt", 'r',encoding='utf-8') as f:
    lines = f.readlines()

# 提取姓名列表
names = []

for line in lines:
    name = line.split(' _ ')[1].strip()  # 使用split()方法分割文本
    names.append(name)

# breakpoint()
for i in tqdm(range(2,max_row+1),desc=str(className)+' 数学数据处理中: ',unit="lines"):
    resoultSheetrow=resoultSheet.max_row+1
    for j in range(0,len(names)):
        if(sourceWorkSheet.cell(row=i,column=markRow_name).value==names[j]):           
            for z in range(1,max_column+1):                
                resoultSheet.cell(row=resoultSheetrow,column=z,value=sourceWorkSheet.cell(row=i,column=z).value)
          
for i in tqdm(range(3,max_row_all+1),desc=str(className)+' 全科数据处理中: ',unit="lines"):
    resoultAllSheetRow=resoultAllSheet.max_row+1   
    for j in range(0,len(names)):
        if(sourceAllWorkSheet.cell(row=i,column=markRow_all_name).value==names[j]):            
            for z in range(1,max_column_all+1):                             
                resoultAllSheet.cell(row=resoultAllSheetRow,column=z,value=sourceAllWorkSheet.cell(row=i,column=z).value)                
                # print(sourceAllWorkSheet.cell(row=i,column=z).value, end=" ") 
                # print(resoultAllSheet.cell(row=resoultAllSheetRow,column=z).value, end=" ")  

# breakpoint()
lastRow=resoultAllSheet.max_row
for i in range(lastRow,0,-1):
    if isinstance(resoultAllSheet.cell(row=i,column=markRow_all_allScore).value,str) or resoultAllSheet.cell(row=i,column=markRow_all_math).value==None:
        lastRow-=1
    else:
        break

# 对数据按照第x列（"总分"列）进行降序排序
sorted_data = sorted(resoultAllSheet.iter_rows(min_row=4, max_row=lastRow, min_col=1, max_col=resoultAllSheet.max_column), key=lambda x: x[markRow_all_allScore-1].value, reverse=True)


# 清除现有数据
resoultAllSheet.delete_rows(4, resoultAllSheet.max_row)

# 将排序后的数据写入新表格
for row in sorted_data:
    resoultAllSheet.append(row)

resoultAllSheet.merge_cells(start_row=1, end_row=1, start_column=1, end_column=max_column_all)
resoultSheet.merge_cells(start_row=1, end_row=1, start_column=1, end_column=max_column)


# 创建一个加粗字体的样式
bold_font = Font(bold=True)
bold_style = NamedStyle(name='bold_style')
bold_style.font = bold_font

# 创建一个填充颜色为红色的PatternFill对象
fill_color = Color(rgb='39c5bb')
fill = PatternFill(fill_type='solid', start_color=fill_color, end_color=fill_color)

fill_color_2 = Color(rgb='dcdcdc')
fill_2 = PatternFill(fill_type='solid', start_color=fill_color_2, end_color=fill_color_2)

fill_color_3 = Color(rgb='7fffd4')
fill_3 = PatternFill(fill_type='solid', start_color=fill_color_3, end_color=fill_color_3)

for row in range(3,resoultAllSheet.max_row+1):
    if not isinstance(resoultAllSheet.cell(row=row,column=markRow_art).value,str):
        for cell in resoultAllSheet[row]:
            cell.fill = fill_3
            
for row in resoultAllSheet.iter_rows(min_row=1, max_row=lastRow, min_col=markRow_all_math, max_col=markRow_all_math):
    for cell in row:
        cell.style = bold_style
        cell.fill = fill
for row in resoultAllSheet.iter_rows(min_row=1, max_row=lastRow, min_col=markRow_all_allScore, max_col=markRow_all_allScore):
    for cell in row:
        cell.style = bold_style
        cell.fill = fill

resoultAllSheet.cell(row=1,column=1).fill=fill_2
resoultAllSheet.cell(row=1,column=1).style=bold_style
resoultAllSheet.cell(row=1,column=1).alignment = Alignment(horizontal='center', vertical='center')

# breakpoint()





resoultExcelFile.save("2023年6月考试 原"+className+"考试成绩.xlsx")


import os
import re
import xml.etree.ElementTree as ET
import xml.dom.minidom as minidom
from tqdm import tqdm
import openpyxl
from openpyxl.styles import Alignment, Font


# 获取当前脚本所在的目录
current_dir = os.path.dirname(os.path.abspath(__file__))

# 列出当前目录下的所有文件
files = [f for f in os.listdir(current_dir) if os.path.isfile(os.path.join(current_dir, f))and f.endswith('.xls')]

# 打印出所有文件名
for f in files:
    # print(f)
    # print(re.findall('\d+', f))
    list=re.findall('\d+', f)
    date=list[0]


# root=ET.Element("root")
doc=minidom.Document()
root = doc.createElement("root")

doc.appendChild(root)

# 创建一个 Workbook 对象
wb = openpyxl.Workbook()

# 删除默认创建的一个 worksheet
default_sheet = wb.active
wb.remove(default_sheet)

# 创建两个 worksheet，分别命名为 Table1 和 Table2
ws1 = wb.create_sheet(title="40")
ws2 = wb.create_sheet(title="100")
ws3 = wb.create_sheet(title="Alldata")

ws3.cell(row=2, column=1, value="考号").font= Font(bold=True)
ws3.cell(row=2, column=2, value="姓名").font= Font(bold=True)
ws3.column_dimensions['A'].width = 13
ws3.freeze_panes = 'B1'

ws1.cell(row=1, column=1, value="姓名").font= Font(bold=True)
ws2.cell(row=1, column=1, value="姓名").font= Font(bold=True)

# 设置字体加粗
font = Font(bold=True)
# 设置水平对齐方式为左对齐
alignment = Alignment(horizontal='left')

# ws1.cell(row=2, column=3).font=font
# ws1.alignment = alignment

cellRowIndex=3
# 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
# 获取文本文件的行数
with open('data155.txt', 'r',encoding="utf-8") as f:
    num_lines = sum(1 for _ in f)

fileName="155"

with open('data'+fileName+'.txt', 'r',encoding='utf-8') as f:
    for line in tqdm(f, total=num_lines,desc='XML文件创建中: ',unit="lines"):
        data=line.strip().split('	')
        
        child_1=doc.createElement("student")              
        child_1.setAttribute("id",data[0])  
        
        child_2_1=doc.createElement("name")
        child_2_1.appendChild(doc.createTextNode(data[1]))
       
        child_2_2=doc.createElement("callCount")
        child_2_2.appendChild(doc.createTextNode("0"))
       
        child_1.appendChild(child_2_1)
        child_1.appendChild(child_2_2)

        root.appendChild(child_1)
        
        ws3.cell(row=cellRowIndex, column=1, value=int(data[0])).alignment = alignment
        ws3.cell(row=cellRowIndex, column=2, value=data[1])
        ws3.cell(row=cellRowIndex, column=1).font=font

        ws1.cell(row=cellRowIndex-1, column=1, value=data[1])
        ws2.cell(row=cellRowIndex-1, column=1, value=data[1])

        cellRowIndex+=1
        
# fileName="142"

# 格式化 XML 字符串
xml_pretty_str = doc.toprettyxml(indent='  ', encoding='utf-8').decode('utf-8')

# 将格式化后的字符串写入文件
with open('data'+fileName+'.xml', 'w', encoding='UTF-8') as f:
    f.write(xml_pretty_str)

wb.save("Exam"+fileName+"Score.xlsx")

